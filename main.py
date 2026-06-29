import openpyxl
from openpyxl.utils import get_column_letter

# All layout candidates with their printable widths in inches
LAYOUTS = [
    ("A4 Portrait",      7.7),
    ("A4 Landscape",    10.7),
    ("A3 Portrait",     11.7),
    ("Legal Landscape", 13.2),
    ("A3 Landscape",    16.5),
]

def col_width_to_inches(excel_width):
    """Convert Excel column width units to inches."""
    return (excel_width * 7 + 5) / 96

def get_total_table_width(ws):
    """Calculate total width of all columns in inches."""
    total = 0
    default_width = ws.sheet_format.defaultColWidth or 8.43

    for col_num in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_num)
        width = ws.column_dimensions[col_letter].width \
                if col_letter in ws.column_dimensions \
                else default_width
        total += col_width_to_inches(width)

    return total

def score_layout(total_width, page_width):
    """
    Score a layout option. Returns -1 if content doesn't fit.
    Higher score = better layout.
    
    Scoring logic:
    - Content must fit within page width (otherwise score = -1)
    - Best score when content fills 75-95% of the page
    - Penalise layouts where content is too small (lots of empty space)
    """
    if total_width > page_width:
        return -1  # Doesn't fit — disqualified

    scale = total_width / page_width

    if scale >= 0.75:
        # Sweet spot: content fills most of the page
        return round(50 + (scale * 50), 1)
    else:
        # Too much empty space — lower score
        return round(scale * 60, 1)

def analyze(filepath):
    """Load an Excel file and print a full layout analysis report."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    total_width = get_total_table_width(ws)

    print(f"\n{'='*48}")
    print(f"  IntelliConverter — Layout Analysis Report")
    print(f"{'='*48}")
    print(f"  Sheet   : {ws.title}")
    print(f"  Columns : {ws.max_column}")
    print(f"  Rows    : {ws.max_row}")
    print(f"  Width   : {total_width:.2f} inches")
    print(f"{'='*48}")
    print(f"  {'Layout':<20} {'Fits':<6} {'Scale':>6}  {'Score':>6}")
    print(f"  {'-'*42}")

    best_name = None
    best_score = -1
    best_scale = 0

    for name, page_width in LAYOUTS:
        fits = total_width <= page_width
        scale = (total_width / page_width) * 100
        score = score_layout(total_width, page_width)

        fits_icon = "✅" if fits else "❌"
        score_display = f"{score:.0f}" if score >= 0 else "—"

        print(f"  {name:<20} {fits_icon:<6} {scale:>5.0f}%  {score_display:>6}")

        if score > best_score:
            best_score = score
            best_name = name
            best_scale = scale

    print(f"  {'-'*42}")
    print(f"\n  ✨ Best layout : {best_name}")
    print(f"  📐 Scale       : {best_scale:.0f}%")
    print(f"  🏆 Score       : {best_score:.0f}/100")
    print(f"\n{'='*48}\n")

if __name__ == "__main__":
    analyze("test.xlsx")